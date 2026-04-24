#!/usr/bin/env python3
"""Export the Firestore `leads` collection to a timestamped Excel workbook.

Usage:
    python scripts/export_leads.py                 # writes ./leads_YYYYMMDD_HHMMSS.xlsx
    python scripts/export_leads.py -o ./out.xlsx   # custom output path

Setup (one-time):
    1. pip install firebase-admin openpyxl
    2. Authenticate with Application Default Credentials (either/or):
         a) Install gcloud SDK and run:
                gcloud auth application-default login
         b) OR download a service-account JSON from
                Firebase Console → Project settings → Service accounts → Generate new private key
            and set:
                export GOOGLE_APPLICATION_CREDENTIALS=/path/to/serviceAccountKey.json

Then run this script from the repo root. The Admin SDK bypasses Firestore
security rules by design, so restricted `read` rules don't block you.
"""
from __future__ import annotations

import argparse
from datetime import datetime, timezone

import firebase_admin
from firebase_admin import credentials, firestore
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ID = "medivoy-1ff3b"

COLUMNS = [
    ("Submitted At (UTC)", 22),
    ("Name", 24),
    ("Email", 32),
    ("Phone", 18),
    ("Source", 10),
    ("Page", 16),
    ("Referrer", 28),
    ("User Agent", 50),
    ("Doc ID", 24),
]


def init_firestore():
    try:
        firebase_admin.get_app()
    except ValueError:
        firebase_admin.initialize_app(options={"projectId": PROJECT_ID})
    return firestore.client()


def fmt_ts(ts) -> str:
    if ts is None:
        return ""
    if hasattr(ts, "astimezone"):
        return ts.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    return str(ts)


def build_workbook(db) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Header row
    header_fill = PatternFill("solid", fgColor="1B3B36")
    header_font = Font(bold=True, color="F7F2EA")
    for col_idx, (label, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    # Rows — ordered by createdAt ascending
    docs = db.collection("leads").order_by("createdAt").stream()
    count = 0
    for doc in docs:
        d = doc.to_dict() or {}
        row = [
            fmt_ts(d.get("createdAt")),
            d.get("name", ""),
            d.get("email", ""),
            d.get("phone", ""),
            d.get("source", ""),
            d.get("page", ""),
            d.get("referrer", ""),
            d.get("userAgent", ""),
            doc.id,
        ]
        ws.append(row)
        count += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(ws.max_row, 1)}"
    return wb, count


def main():
    parser = argparse.ArgumentParser(description="Export Medivoyage leads to xlsx.")
    parser.add_argument(
        "-o", "--output",
        default=None,
        help="Output .xlsx path (default: leads_YYYYMMDD_HHMMSS.xlsx in cwd)",
    )
    args = parser.parse_args()

    out_path = args.output or f"leads_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    db = init_firestore()
    wb, count = build_workbook(db)
    wb.save(out_path)
    print(f"Exported {count} lead(s) → {out_path}")


if __name__ == "__main__":
    main()
